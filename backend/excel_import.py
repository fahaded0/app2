"""Excel (.xlsx) import for items. Matching order: barcode -> internal_code -> name -> manual_review."""
import hashlib
import io
import os
from typing import Optional

_EXCEL_TXN_HOOKS_ACTIVE = (
    os.environ.get("APP_ENV") == "test" and
    os.environ.get("TRANSACTION_TEST_HOOKS_ENABLED") == "true"
)


class ExcelTestFailure(Exception):
    """Raised by _excel_check_fail_point during test-injected transactional failures."""


class ExcelWriteConflict(Exception):
    """Raised when a row transaction fails with a write conflict after one retry."""


class ExcelCASConflict(Exception):
    """Raised when stock_entries CAS update matches zero rows (concurrent modification)."""


def _excel_check_fail_point(fail_after, stage: str) -> None:
    if _EXCEL_TXN_HOOKS_ACTIVE and fail_after == stage:
        raise ExcelTestFailure(f"[test] Forced failure at {stage}")


def _validate_excel_replay(
    prior: dict,
    *,
    department_id: str,
    item_id: str,
    new_balance: int,
) -> None:
    """Validate a prior ledger record is genuinely this row's replay.
    Raises ExcelWriteConflict if any field mismatches — prevents key-namespace collisions.
    item_id is mandatory and always compared against the prior record.
    """
    checks = [
        ("schema_version", 2),
        ("source", "excel_import"),
        ("entry_type", "physical_count"),
        ("department_id", department_id),
        ("item_id", item_id),
        ("new_balance", new_balance),
    ]
    mismatches = [f for f, exp in checks if prior.get(f) != exp]
    if mismatches:
        raise ExcelWriteConflict(
            f"Idempotency key reused with mismatched payload in Excel import: {mismatches}"
        )


async def _resolve_item_id_for_replay(db, entry: dict, session=None) -> str:
    """Resolve the expected item_id before calling _validate_excel_replay.

    Resolution order: existing_id → internal_code lookup → barcode lookup → conflict.
    If lookup fails, the idempotency key was reused for a different item — raise ExcelWriteConflict.
    """
    if entry.get("existing_id"):
        return entry["existing_id"]
    if entry.get("internal_code"):
        doc = await db.items.find_one(
            {"internal_code": entry["internal_code"]}, {"_id": 0}, session=session
        )
        if doc:
            return doc["id"]
    if entry.get("barcode"):
        doc = await db.items.find_one(
            {"barcode": entry["barcode"]}, {"_id": 0}, session=session
        )
        if doc:
            return doc["id"]
    raise ExcelWriteConflict(
        "Idempotency key reused but expected item cannot be identified — refusing to skip"
    )

from openpyxl import load_workbook
from motor.motor_asyncio import AsyncIOMotorDatabase
from pymongo.errors import DuplicateKeyError

from models import _new_id, _now_iso
import ledger as ledger_mod

# Expected headers (lowercased). All optional except internal_code OR barcode.
EXPECTED_HEADERS = [
    "internal_code", "barcode", "name", "category", "unit",
    "min_level", "critical_threshold", "max_level",
    "department_code", "balance",
]


def _norm(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_int(v, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def parse_workbook(file_bytes: bytes) -> list[dict]:
    """Read the first worksheet and return a list of row dicts keyed by header."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers_raw = [(_norm(h) or "").lower() for h in rows[0]]
    parsed: list[dict] = []
    for i, row in enumerate(rows[1:], start=2):     # 1-based excel row, header on row 1
        if all(c is None or c == "" for c in row):
            continue
        item: dict = {"__row__": i}
        for h, c in zip(headers_raw, row):
            if h:
                item[h] = c
        parsed.append(item)
    return parsed


async def _find_match(db: AsyncIOMotorDatabase, row: dict) -> tuple[Optional[dict], str]:
    """Return (existing_item or None, match_strategy).
    Strategies: 'barcode', 'internal_code', 'name', 'none'."""
    barcode = _norm(row.get("barcode"))
    if barcode:
        existing = await db.items.find_one({"barcode": barcode}, {"_id": 0})
        if existing:
            return existing, "barcode"
    code = _norm(row.get("internal_code"))
    if code:
        existing = await db.items.find_one({"internal_code": code}, {"_id": 0})
        if existing:
            return existing, "internal_code"
    name = _norm(row.get("name"))
    if name:
        existing = await db.items.find_one(
            {"$or": [{"name_en": name}, {"name_ar": name}]}, {"_id": 0}
        )
        if existing:
            return existing, "name"
    return None, "none"


async def analyse(db: AsyncIOMotorDatabase, file_bytes: bytes) -> dict:
    """Build a preview report describing what a commit would do."""
    rows = parse_workbook(file_bytes)
    departments = {d["code"]: d for d in await db.departments.find({}, {"_id": 0}).to_list(500)}

    summary = {
        "total_rows": len(rows),
        "to_create": [],
        "to_update": [],
        "manual_review": [],
        "errors": [],
    }
    for r in rows:
        ref = r["__row__"]
        name = _norm(r.get("name"))
        code = _norm(r.get("internal_code"))
        barcode = _norm(r.get("barcode"))

        if not (name or code or barcode):
            summary["errors"].append({"row": ref, "reason": "Empty key fields (need at least one of internal_code/barcode/name)"})
            continue

        existing, strategy = await _find_match(db, r)
        dept_code = _norm(r.get("department_code"))
        if dept_code and dept_code not in departments:
            summary["errors"].append({"row": ref, "reason": f"Unknown department_code '{dept_code}'"})
            continue

        entry = {
            "row": ref,
            "strategy": strategy,
            "existing_id": existing["id"] if existing else None,
            "internal_code": code,
            "barcode": barcode,
            "name": name,
            "category": _norm(r.get("category")),
            "unit": _norm(r.get("unit")),
            "min_level": _to_int(r.get("min_level")),
            "critical_threshold": _to_int(r.get("critical_threshold")),
            "max_level": _to_int(r.get("max_level")),
            "department_code": dept_code,
            "balance": _to_int(r.get("balance"), default=None) if r.get("balance") not in (None, "") else None,
        }
        if existing:
            summary["to_update"].append(entry)
        elif strategy == "none" and not code:
            # No code and no match → needs manual review
            summary["manual_review"].append(entry)
        else:
            summary["to_create"].append(entry)
    return summary


def _calc_status(balance: int, min_level: int, critical: int) -> str:
    if balance == 0:
        return "zero_level"
    if balance < critical:
        return "critical_level"
    return "available"


async def commit(
    db: AsyncIOMotorDatabase,
    file_bytes: bytes,
    user: dict,
    *,
    include_manual_review: bool = False,
    client=None,
    audit_callback=None,
    fail_after=None,
) -> dict:
    """Apply the import. Returns a summary of created/updated counts."""
    workbook_sha = hashlib.sha256(file_bytes).hexdigest()
    plan = await analyse(db, file_bytes)
    departments = {d["code"]: d for d in await db.departments.find({}, {"_id": 0}).to_list(500)}
    created = 0
    updated = 0
    stock_updated = 0
    skipped = 0

    rows_to_process = plan["to_create"] + plan["to_update"]
    if include_manual_review:
        rows_to_process += plan["manual_review"]

    for entry in rows_to_process:
        row_number = entry.get("row", 0)

        # For rows with stock updates, use a per-row transaction with idempotency guard
        if entry["balance"] is not None and entry["department_code"]:
            dept = departments.get(entry["department_code"])
            if not dept:
                skipped += 1
                continue

            if client is None:
                raise RuntimeError("client is required for stock-changing Excel rows")

            idem_key = ledger_mod.workbook_row_idempotency_key(file_bytes, row_number)

            # Pre-transaction idempotency check — skip row if already processed
            prior = await db.stock_transactions.find_one(
                {"idempotency_key": idem_key, "schema_version": 2}, {"_id": 0}
            )
            if prior:
                expected_item_id = await _resolve_item_id_for_replay(db, entry)
                _validate_excel_replay(prior, department_id=dept["id"],
                                       item_id=expected_item_id, new_balance=entry["balance"])
                skipped += 1
                continue

            # Run inside a MongoDB transaction
            async def _row_callback(session, _entry=entry, _dept=dept, _idem_key=idem_key):
                return await _process_row(
                    db, file_bytes, user, _entry, _dept, _idem_key,
                    session=session, audit_callback=audit_callback, fail_after=fail_after,
                )
            async with await client.start_session() as session:
                try:
                    row_result = await session.with_transaction(_row_callback)
                except (DuplicateKeyError, ExcelCASConflict):
                    prior = await db.stock_transactions.find_one(
                        {"idempotency_key": idem_key, "schema_version": 2}, {"_id": 0}
                    )
                    if prior:
                        expected_item_id = await _resolve_item_id_for_replay(db, entry)
                        _validate_excel_replay(prior, department_id=dept["id"],
                                               item_id=expected_item_id, new_balance=entry["balance"])
                        skipped += 1
                        continue
                    # Baseline-race: retry once
                    try:
                        async with await client.start_session() as _s2:
                            row_result = await _s2.with_transaction(_row_callback)
                    except (DuplicateKeyError, ExcelCASConflict):
                        prior2 = await db.stock_transactions.find_one(
                            {"idempotency_key": idem_key, "schema_version": 2}, {"_id": 0}
                        )
                        if prior2:
                            expected_item_id = await _resolve_item_id_for_replay(db, entry)
                            _validate_excel_replay(prior2, department_id=dept["id"],
                                                   item_id=expected_item_id, new_balance=entry["balance"])
                            skipped += 1
                            continue
                        raise ExcelWriteConflict(
                            "Write conflict on Excel import row. Please retry."
                        )

            if row_result["item_created"]:
                created += 1
            elif row_result["item_updated"]:
                updated += 1
            stock_updated += 1
        else:
            # No stock write — item-only create/update outside a transaction
            item_result = await _upsert_item(db, entry)
            if item_result == "created":
                created += 1
            elif item_result == "updated":
                updated += 1
            elif item_result == "skipped":
                skipped += 1

    return {
        "created_items": created,
        "updated_items": updated,
        "stock_entries_touched": stock_updated,
        "skipped": skipped + len(plan["errors"]),
        "errors": plan["errors"],
        "manual_review_count": 0 if include_manual_review else len(plan["manual_review"]),
        "workbook_sha256": workbook_sha,
    }


async def _upsert_item(db: AsyncIOMotorDatabase, entry: dict, session=None) -> str:
    """Create or update the item record. Returns 'created', 'updated', or 'skipped'."""
    if entry["existing_id"]:
        update_doc = {k: v for k, v in {
            "barcode": entry["barcode"],
            "name_en": entry["name"],
            "name_ar": entry["name"],
            "category": entry["category"],
            "unit": entry["unit"],
            "min_level": entry["min_level"],
            "critical_threshold": entry["critical_threshold"],
            "max_level": entry["max_level"],
            "updated_at": _now_iso(),
        }.items() if v not in (None, "")}
        if update_doc:
            await db.items.update_one({"id": entry["existing_id"]}, {"$set": update_doc}, session=session)
            return "updated"
        return "skipped"
    else:
        if not entry["internal_code"]:
            return "skipped"
        item_id = _new_id()
        await db.items.insert_one({
            "id": item_id,
            "internal_code": entry["internal_code"],
            "barcode": entry["barcode"],
            "udi": None,
            "gtin": None,
            "name_ar": entry["name"] or entry["internal_code"],
            "name_en": entry["name"] or entry["internal_code"],
            "category": entry["category"] or "Other",
            "unit": entry["unit"] or "PCS",
            "min_level": entry["min_level"],
            "critical_threshold": entry["critical_threshold"],
            "max_level": entry["max_level"],
            "reorder_qty": 0,
            "lead_time_days": 0,
            "alternative_item_id": None,
            "is_life_saving": False,
            "is_crash_cart": False,
            "requires_expiry": False,
            "supplier": None,
            "is_active": True,
            "notes": None,
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        }, session=session)
        return "created"


async def _process_row(
    db: AsyncIOMotorDatabase,
    file_bytes: bytes,
    user: dict,
    entry: dict,
    dept: dict,
    idem_key: str,
    session=None,
    audit_callback=None,
    fail_after=None,
) -> dict:
    """Process one excel row inside an optional transaction. Returns item_created/item_updated flags."""
    # Idempotency recheck inside the transaction
    prior = await db.stock_transactions.find_one(
        {"idempotency_key": idem_key, "schema_version": 2}, {"_id": 0}, session=session
    )
    if prior:
        expected_item_id = await _resolve_item_id_for_replay(db, entry, session=session)
        _validate_excel_replay(prior, department_id=dept["id"],
                               item_id=expected_item_id, new_balance=entry["balance"])
        return {"item_created": False, "item_updated": False}

    # Upsert item (create or update)
    item_result = await _upsert_item(db, entry, session=session)
    item_id = entry["existing_id"] or (
        (await db.items.find_one({"internal_code": entry["internal_code"]}, {"_id": 0}, session=session) or {}).get("id")
    )
    if not item_id:
        return {"item_created": False, "item_updated": False}

    item_doc = await db.items.find_one({"id": item_id}, {"_id": 0}, session=session)
    if not item_doc:
        return {"item_created": False, "item_updated": False}

    existing_stock = await db.stock_entries.find_one(
        {"department_id": dept["id"], "item_id": item_id}, session=session
    )
    previous_balance = existing_stock["balance"] if existing_stock else 0
    quantity_change = entry["balance"] - previous_balance
    new_balance = entry["balance"]
    status = _calc_status(new_balance, item_doc["min_level"], item_doc["critical_threshold"])

    stock_id = existing_stock["id"] if existing_stock else _new_id()
    previous_lv = existing_stock.get("ledger_version", 0) if existing_stock else 0

    if existing_stock is None:
        seq_no = 1
    elif previous_lv == 0:
        baseline_status = _calc_status(previous_balance, item_doc["min_level"], item_doc["critical_threshold"])
        await ledger_mod.ensure_v2_baseline(
            db,
            department_id=dept["id"],
            item_id=item_id,
            entry_id=stock_id,
            balance=previous_balance,
            user_id=user["id"],
            user_name=user["full_name"],
            idempotency_key=f"baseline:{dept['id']}:{item_id}",
            status=baseline_status,
            source="ledger_v2_cutover",
            session=session,
        )
        seq_no = 2
    else:
        seq_no = previous_lv + 1

    stock_doc = {
        "department_id": dept["id"],
        "item_id": item_id,
        "balance": new_balance,
        "status": status,
        "last_updated_by": user["id"],
        "last_updated_by_name": user["full_name"],
        "last_updated_at": _now_iso(),
        "shortage_start": _now_iso() if status in ("zero_level", "critical_level") else None,
        "notes": "Excel import",
    }
    if existing_stock:
        filter_doc = {"id": stock_id, "balance": previous_balance}
        if previous_lv >= 1:
            filter_doc["ledger_version"] = previous_lv
        else:
            filter_doc["$or"] = [{"ledger_version": {"$exists": False}}, {"ledger_version": 0}]
        upd = await db.stock_entries.update_one(
            filter_doc, {"$set": {**stock_doc, "ledger_version": seq_no}}, session=session
        )
        if upd.matched_count != 1:
            raise ExcelCASConflict("Concurrent modification on stock entry during Excel import: please retry.")
    else:
        await db.stock_entries.insert_one({"id": stock_id, **stock_doc, "ledger_version": seq_no}, session=session)

    _excel_check_fail_point(fail_after, "excel_stock_update")

    ledger_entry = ledger_mod.build_ledger_entry(
        department_id=dept["id"],
        item_id=item_id,
        entry_type="physical_count",
        sequence_no=seq_no,
        previous_balance=previous_balance,
        quantity_change=quantity_change,
        new_balance=new_balance,
        user_id=user["id"],
        user_name=user["full_name"],
        actor_type="user",
        source="excel_import",
        idempotency_key=idem_key,
        status=status,
        entry_id=stock_id,
    )
    await ledger_mod.insert_ledger_entry(db, ledger_entry, session=session)

    if audit_callback:
        await audit_callback(session=session, item_id=item_id, dept_id=dept["id"],
                             entry=ledger_entry)

    return {"item_created": item_result == "created", "item_updated": item_result == "updated"}
