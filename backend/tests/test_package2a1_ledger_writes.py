"""Package 2A-1 — Immutable ledger v2 integration tests.

Tests cover all stock-write paths that produce ledger entries:
  upsert_stock, receive_request, excel_import.commit, seed._ensure_sample_stock

All tests use unique item codes / idempotency keys to remain independent.
Credentials are read from environment variables with backward-compatible defaults.
"""
import hashlib
import io
import os
import pathlib
import sys
import uuid
import pytest
import requests

# Allow direct import of backend modules for unit tests (no network/DB required)
sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").strip().rstrip("/")
API = f"{BASE_URL}/api"

_ADMIN_EMAIL = os.environ.get("TEST_ADMIN_EMAIL", "admin@medstock.sa")
_ADMIN_PW    = os.environ.get("TEST_ADMIN_PASSWORD", "Admin@12345")

# Aliases used by newer test classes
ADMIN_EMAIL = _ADMIN_EMAIL
ADMIN_PASSWORD = _ADMIN_PW


def _new_id() -> str:
    return uuid.uuid4().hex


# ---------------------------------------------------------------------------
# Auth / HTTP helpers
# ---------------------------------------------------------------------------

def _login(email: str, pw: str) -> str:
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": pw}, timeout=15)
    assert r.status_code == 200, f"Login failed for {email}: {r.status_code} {r.text}"
    return r.json()["access_token"]


def _h(token: str, extra: dict | None = None) -> dict:
    headers = {"Authorization": f"Bearer {token}"}
    if extra:
        headers.update(extra)
    return headers


@pytest.fixture(scope="module")
def admin_token():
    return _login(_ADMIN_EMAIL, _ADMIN_PW)


# ---------------------------------------------------------------------------
# Setup helpers
# ---------------------------------------------------------------------------

def _unique(prefix: str = "") -> str:
    return f"{prefix}{uuid.uuid4().hex[:10].upper()}"


def _create_item(token: str, internal_code: str) -> dict:
    payload = {
        "internal_code": internal_code,
        "name_ar": f"عنصر اختبار {internal_code}",
        "name_en": f"Ledger Test Item {internal_code}",
        "category": "Other",
        "unit": "PCS",
        "min_level": 20,
        "critical_threshold": 10,
        "max_level": 200,
    }
    r = requests.post(f"{API}/items", headers=_h(token), json=payload, timeout=15)
    assert r.status_code == 200, f"create item failed: {r.status_code} {r.text}"
    return r.json()


def _get_er_dept(token: str) -> dict:
    depts = requests.get(f"{API}/departments", headers=_h(token)).json()
    return next(d for d in depts if d["code"] == "ER")


def _upsert_stock(token: str, item_id: str, dept_id: str, balance: int,
                  idem_key: str | None = None) -> requests.Response:
    headers = _h(token)
    if idem_key:
        headers["Idempotency-Key"] = idem_key
    return requests.post(f"{API}/stock", headers=headers, json={
        "department_id": dept_id,
        "item_id": item_id,
        "balance": balance,
        "notes": "ledger test",
    }, timeout=15)


def _get_ledger(token: str, item_id: str, dept_id: str) -> list[dict]:
    """Fetch v2 ledger entries for a specific item+dept pair via GET /stock/transactions."""
    r = requests.get(
        f"{API}/stock/transactions",
        headers=_h(token),
        params={"item_id": item_id, "department_id": dept_id},
        timeout=15,
    )
    if r.status_code == 404:
        return []
    assert r.status_code == 200, f"get_ledger failed: {r.status_code} {r.text}"
    all_txns = r.json() if isinstance(r.json(), list) else r.json().get("transactions", r.json())
    return [t for t in all_txns if t.get("schema_version") == 2]


def _create_and_seed(token: str, balance: int = 100) -> dict:
    """Create a unique item and seed its stock balance. Returns {item, dept, initial_balance}."""
    item = _create_item(token, _unique("L"))
    dept = _get_er_dept(token)
    r = _upsert_stock(token, item["id"], dept["id"], balance)
    assert r.status_code == 200, f"seed upsert failed: {r.text}"
    return {"item": item, "dept": dept, "initial_balance": balance}


def _get_request(token: str, req_id: str) -> dict:
    r = requests.get(f"{API}/requests", headers=_h(token), timeout=15)
    assert r.status_code == 200, f"get_requests failed: {r.status_code} {r.text}"
    data = r.json()
    items_list = data if isinstance(data, list) else data.get("requests", data.get("items", []))
    req = next((x for x in items_list if x.get("id") == req_id), None)
    assert req is not None, f"request {req_id} not found in list"
    return req


def _create_request(token, dept_id, item_id, qty=20):
    r = requests.post(f"{API}/requests", headers=_h(token), json={
        "department_id": dept_id,
        "item_id": item_id,
        "requested_qty": qty,
        "priority": "routine",
        "reason": "ledger receive test",
    }, timeout=15)
    assert r.status_code == 200, f"create request failed: {r.text}"
    return r.json()


def _approve_request(token, req_id, qty=20):
    r = requests.post(f"{API}/requests/{req_id}/approve", headers=_h(token),
                      json={"approved_qty": qty}, timeout=15)
    assert r.status_code == 200, f"approve failed: {r.text}"


def _dispatch_request(token, req_id, qty=20):
    r = requests.post(f"{API}/requests/{req_id}/dispatch", headers=_h(token),
                      json={"dispatched_qty": qty, "backorder": False}, timeout=15)
    assert r.status_code == 200, f"dispatch failed: {r.text}"


# ---------------------------------------------------------------------------
# Non-integration unit tests — no DB, no network
# ---------------------------------------------------------------------------

class TestBuildLedgerEntryProtection:
    """Unit tests — no DB, no network."""

    def test_reserved_field_override_rejected(self):
        from ledger import build_ledger_entry
        with pytest.raises(ValueError, match="reserved"):
            build_ledger_entry(
                department_id="d1", item_id="i1", entry_type="adjustment",
                sequence_no=2, previous_balance=10, quantity_change=-5, new_balance=5,
                user_id="u1", user_name="User", actor_type="user",
                source="test", idempotency_key="ik1", status="available", entry_id="e1",
                # Reserved field override attempt:
                schema_version=99,
            )

    def test_id_override_rejected(self):
        from ledger import build_ledger_entry
        with pytest.raises(ValueError, match="reserved"):
            build_ledger_entry(
                department_id="d1", item_id="i1", entry_type="adjustment",
                sequence_no=1, previous_balance=10, quantity_change=5, new_balance=15,
                user_id="u1", user_name="User", source="test",
                idempotency_key="idem-2", status="available", entry_id="e1",
                id="evil-override",
            )

    def test_created_at_override_rejected(self):
        from ledger import build_ledger_entry
        with pytest.raises(ValueError, match="reserved"):
            build_ledger_entry(
                department_id="d1", item_id="i1", entry_type="adjustment",
                sequence_no=1, previous_balance=10, quantity_change=5, new_balance=15,
                user_id="u1", user_name="User", source="test",
                idempotency_key="idem-3", status="available", entry_id="e1",
                created_at="evil-time",
            )

    def test_non_reserved_extra_allowed(self):
        from ledger import build_ledger_entry
        doc = build_ledger_entry(
            department_id="d1", item_id="i1", entry_type="adjustment",
            sequence_no=2, previous_balance=10, quantity_change=-5, new_balance=5,
            user_id="u1", user_name="User", actor_type="user",
            source="test", idempotency_key="ik3", status="available", entry_id="e1",
            # Non-reserved extras:
            custom_field="value", request_id="req1",
        )
        assert doc["custom_field"] == "value"
        assert doc["request_id"] == "req1"
        assert doc["schema_version"] == 2  # not overridden

    def test_balance_invariant_enforced(self):
        from ledger import build_ledger_entry
        with pytest.raises(ValueError, match="integrity"):
            build_ledger_entry(
                department_id="d1", item_id="i1", entry_type="adjustment",
                sequence_no=1, previous_balance=10, quantity_change=5, new_balance=20,  # 10+5 != 20
                user_id="u1", user_name="User",
                source="test", idempotency_key="ik4", status="available", entry_id="e1",
            )

    def test_transaction_id_used_as_doc_id(self):
        from ledger import build_ledger_entry
        doc = build_ledger_entry(
            department_id="d1", item_id="i1", entry_type="issue",
            sequence_no=1, previous_balance=10, quantity_change=-5, new_balance=5,
            user_id="u1", user_name="User",
            source="test", idempotency_key="ik5", status="available", entry_id="e1",
            transaction_id="my-explicit-id",
        )
        assert doc["id"] == "my-explicit-id"


# ---------------------------------------------------------------------------
# Test 1 — upsert_stock creates a ledger entry on first write
# ---------------------------------------------------------------------------

class TestUpsertStockCreatesLedger:
    pytestmark = pytest.mark.integration

    def test_ledger_entry_created(self, admin_token):
        item = _create_item(admin_token, _unique("T1U"))
        dept = _get_er_dept(admin_token)
        idem_key = _unique("IK1")

        r = _upsert_stock(admin_token, item["id"], dept["id"], balance=50, idem_key=idem_key)
        assert r.status_code == 200, r.text

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        assert len(entries) >= 1, "expected at least one ledger entry"

        entry = entries[0]
        assert entry["schema_version"] == 2
        assert entry["department_id"] == dept["id"]
        assert entry["item_id"] == item["id"]
        assert entry["new_balance"] == 50
        assert entry["previous_balance"] == 0
        assert entry["quantity_change"] == 50
        assert entry["quantity_change"] == entry["delta"]
        assert entry["entry_type"] == "opening_balance"
        assert entry["source"] == "upsert_stock"
        assert entry["actor_type"] == "user"
        assert entry["sequence_no"] == 1

    def test_balance_invariant(self, admin_token):
        item = _create_item(admin_token, _unique("T1B"))
        dept = _get_er_dept(admin_token)
        r = _upsert_stock(admin_token, item["id"], dept["id"], balance=75)
        assert r.status_code == 200, r.text

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        assert entries, "no ledger entries found"
        for e in entries:
            assert e["new_balance"] == e["previous_balance"] + e["quantity_change"], (
                f"balance invariant violated: {e['previous_balance']} + {e['quantity_change']} != {e['new_balance']}"
            )


# ---------------------------------------------------------------------------
# Test 2 — second upsert creates a second ledger entry with incremented sequence_no
# ---------------------------------------------------------------------------

class TestUpsertStockSequenceIncrement:
    pytestmark = pytest.mark.integration

    def test_sequence_increments(self, admin_token):
        ctx = _create_and_seed(admin_token, balance=100)
        item_id = ctx["item"]["id"]
        dept_id = ctx["dept"]["id"]

        r = _upsert_stock(admin_token, item_id, dept_id, balance=80)
        assert r.status_code == 200, r.text

        entries = _get_ledger(admin_token, item_id, dept_id)
        seq_nos = sorted(e["sequence_no"] for e in entries)
        assert seq_nos == list(range(1, len(seq_nos) + 1)), f"sequence gap: {seq_nos}"
        assert len(seq_nos) >= 2

    def test_adjustment_entry_type_on_subsequent_write(self, admin_token):
        ctx = _create_and_seed(admin_token, balance=100)
        item_id = ctx["item"]["id"]
        dept_id = ctx["dept"]["id"]

        r = _upsert_stock(admin_token, item_id, dept_id, balance=60)
        assert r.status_code == 200, r.text

        entries = sorted(_get_ledger(admin_token, item_id, dept_id), key=lambda e: e["sequence_no"])
        assert entries[0]["entry_type"] == "opening_balance"
        assert entries[1]["entry_type"] == "adjustment"


# ---------------------------------------------------------------------------
# Test 3 — idempotent replay: same Idempotency-Key does not create a second ledger entry
# ---------------------------------------------------------------------------

class TestUpsertStockIdempotency:
    pytestmark = pytest.mark.integration

    def test_no_duplicate_ledger_entry(self, admin_token):
        item = _create_item(admin_token, _unique("T3I"))
        dept = _get_er_dept(admin_token)
        idem_key = _unique("IK3")

        r1 = _upsert_stock(admin_token, item["id"], dept["id"], balance=50, idem_key=idem_key)
        assert r1.status_code == 200, r1.text

        r2 = _upsert_stock(admin_token, item["id"], dept["id"], balance=50, idem_key=idem_key)
        assert r2.status_code == 200, r2.text
        assert r2.json().get("idempotent_replay") is True

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        assert len(entries) == 1, f"expected 1 ledger entry after replay, got {len(entries)}"


# ---------------------------------------------------------------------------
# Test 4 — upsert_stock rollback via test hook: no ledger entry created
# ---------------------------------------------------------------------------

class TestUpsertStockRollback:
    pytestmark = pytest.mark.integration

    def _assert_upsert_rollback(self, admin_token, item, dept, idem_key):
        """Assert that a failed upsert_stock left absolutely no residue."""
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        assert entries == [], f"expected no ledger entries after rollback, got {entries}"

        async def _check():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                alert = await db.alerts.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                audit = await db.audit_logs.find_one(
                    {"action": "upsert_stock", "new_value.idempotency_key": idem_key}, {"_id": 0}
                )
                return se, alert, audit
            finally:
                await c.close()
        se, alert, audit = asyncio.run(_check())
        assert se is None, f"stock entry must not exist after rollback: {se}"
        assert alert is None, f"alert must not exist after rollback: {alert}"
        assert audit is None, f"upsert_stock audit must not exist after rollback: {audit}"

    def test_rollback_at_ledger_insert_leaves_no_entry(self, admin_token):
        item = _create_item(admin_token, _unique("T4R"))
        dept = _get_er_dept(admin_token)
        idem_key = _unique("T4RK")

        r = requests.post(f"{API}/stock",
                          headers=_h(admin_token, {"X-Test-Txn-Fail-After": "ledger_insert",
                                                   "Idempotency-Key": idem_key}),
                          json={"department_id": dept["id"], "item_id": item["id"],
                                "balance": 50, "notes": "rollback test"}, timeout=15)
        assert r.status_code == 503, f"expected 503, got {r.status_code}: {r.text}"
        self._assert_upsert_rollback(admin_token, item, dept, idem_key)

    def test_rollback_at_stock_update_leaves_no_entry(self, admin_token):
        item = _create_item(admin_token, _unique("T4S"))
        dept = _get_er_dept(admin_token)
        idem_key = _unique("T4SK")

        r = requests.post(f"{API}/stock",
                          headers=_h(admin_token, {"X-Test-Txn-Fail-After": "stock_update",
                                                   "Idempotency-Key": idem_key}),
                          json={"department_id": dept["id"], "item_id": item["id"],
                                "balance": 50, "notes": "rollback stock_update test"}, timeout=15)
        assert r.status_code == 503, f"expected 503, got {r.status_code}: {r.text}"
        self._assert_upsert_rollback(admin_token, item, dept, idem_key)


# ---------------------------------------------------------------------------
# Test 4b — receive_request rollback via test hook
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestReceiveRollback:

    def test_receive_rollback_leaves_no_entry(self, admin_token):
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        item = _create_item(admin_token, _unique("T_RVRB"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], 0)
        req = _create_request(admin_token, dept["id"], item["id"], qty=20)
        _approve_request(admin_token, req["id"], qty=20)
        _dispatch_request(admin_token, req["id"], qty=20)
        idem_key = _unique("RVRBK")

        async def _before():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                req_doc = await db.stock_requests.find_one({"id": req["id"]}, {"_id": 0})
                txn_count = await db.stock_transactions.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"], "schema_version": 2}
                )
                audit_count = await db.audit_logs.count_documents(
                    {"action": "receive_request", "new_value.idempotency_key": idem_key}
                )
                alert_count = await db.alerts.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"]}
                )
                return se, req_doc, txn_count, audit_count, alert_count
            finally:
                await c.close()
        se_before, req_before, ledger_count_before, audit_count_before, alert_count_before = asyncio.run(_before())

        r = requests.post(
            f"{API}/requests/{req['id']}/receive",
            headers=_h(admin_token, {"X-Test-Txn-Fail-After": "ledger_insert",
                                     "Idempotency-Key": idem_key}),
            json={"received_qty": 20}, timeout=15,
        )
        assert r.status_code == 503, f"expected 503, got {r.status_code}: {r.text}"

        async def _after():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                req_doc = await db.stock_requests.find_one({"id": req["id"]}, {"_id": 0})
                txn_count = await db.stock_transactions.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"], "schema_version": 2}
                )
                receive_entry = await db.stock_transactions.find_one(
                    {"idempotency_key": idem_key, "schema_version": 2}, {"_id": 0}
                )
                audit = await db.audit_logs.find_one(
                    {"action": "receive_request", "new_value.idempotency_key": idem_key}, {"_id": 0}
                )
                alert_count = await db.alerts.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"]}
                )
                return se, req_doc, txn_count, receive_entry, audit, alert_count
            finally:
                await c.close()
        se_after, req_after, ledger_count_after, receive_entry, audit, alert_count_after = asyncio.run(_after())

        bal_before = se_before["balance"] if se_before else 0
        bal_after = se_after["balance"] if se_after else 0
        assert bal_after == bal_before, f"stock balance changed: {bal_before} → {bal_after}"

        lv_before = se_before.get("ledger_version", 0) if se_before else 0
        lv_after = se_after.get("ledger_version", 0) if se_after else 0
        assert lv_after == lv_before, f"ledger_version changed: {lv_before} → {lv_after}"

        assert ledger_count_after == ledger_count_before, \
            f"ledger count changed: {ledger_count_before} → {ledger_count_after}"
        assert receive_entry is None, \
            f"receive ledger record must not exist after rollback: {receive_entry}"
        assert audit is None, \
            f"receive_request audit must not exist after rollback: {audit}"
        assert alert_count_after == alert_count_before, \
            f"alert count changed: {alert_count_before} → {alert_count_after}"
        assert req_after.get("received_qty", 0) == req_before.get("received_qty", 0), \
            f"received_qty changed: {req_before.get('received_qty')} → {req_after.get('received_qty')}"
        assert req_after.get("status") == req_before.get("status"), \
            f"request status changed: {req_before.get('status')} → {req_after.get('status')}"


# ---------------------------------------------------------------------------
# Test 5 — receive_request creates a ledger entry
# ---------------------------------------------------------------------------

class TestReceiveRequestLedger:
    pytestmark = pytest.mark.integration

    def _setup_request(self, token: str) -> dict:
        """Create item, seed stock, create a supply request, approve and dispatch it."""
        item = _create_item(token, _unique("T5R"))
        dept = _get_er_dept(token)
        _upsert_stock(token, item["id"], dept["id"], balance=0)

        req = _create_request(token, dept["id"], item["id"], qty=20)
        _approve_request(token, req["id"], qty=20)
        _dispatch_request(token, req["id"], qty=20)

        return {"req": req, "item": item, "dept": dept}

    def test_receive_creates_ledger_entry(self, admin_token):
        ctx = self._setup_request(admin_token)
        req_id = ctx["req"]["id"]
        item_id = ctx["item"]["id"]
        dept_id = ctx["dept"]["id"]

        entries_before = _get_ledger(admin_token, item_id, dept_id)

        r = requests.post(f"{API}/requests/{req_id}/receive", headers=_h(admin_token),
                          json={"received_qty": 20}, timeout=15)
        assert r.status_code == 200, r.text

        entries_after = _get_ledger(admin_token, item_id, dept_id)
        new_entries = [e for e in entries_after if e not in entries_before]
        assert len(new_entries) == 1, f"expected 1 new ledger entry, got {len(new_entries)}"

        entry = new_entries[0]
        assert entry["entry_type"] == "receive"
        assert entry["quantity_change"] == 20
        assert entry["new_balance"] == 20
        assert entry["source"] == "receive_request"
        assert entry.get("request_id") == req_id

    def test_receive_idempotent_replay(self, admin_token):
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        ctx = self._setup_request(admin_token)
        req_id = ctx["req"]["id"]
        item_id = ctx["item"]["id"]
        dept_id = ctx["dept"]["id"]
        idem_key = _unique("IK5R")

        r1 = requests.post(f"{API}/requests/{req_id}/receive",
                           headers=_h(admin_token, {"Idempotency-Key": idem_key}),
                           json={"received_qty": 20}, timeout=15)
        assert r1.status_code == 200, r1.text

        # Capture state after first receive (before replay)
        async def _capture_state():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                req_doc = await db.stock_requests.find_one({"id": req_id}, {"_id": 0})
                se = await db.stock_entries.find_one(
                    {"department_id": dept_id, "item_id": item_id}, {"_id": 0}
                )
                receive_count = await db.stock_transactions.count_documents(
                    {"department_id": dept_id, "item_id": item_id,
                     "schema_version": 2, "entry_type": "receive"}
                )
                audit_count = await db.audit_logs.count_documents(
                    {"action": "receive_request", "new_value.idempotency_key": idem_key}
                )
                return req_doc, se, receive_count, audit_count
            finally:
                await c.close()
        req_before, se_before, receive_count_before, audit_count_before = asyncio.run(_capture_state())

        r2 = requests.post(f"{API}/requests/{req_id}/receive",
                           headers=_h(admin_token, {"Idempotency-Key": idem_key}),
                           json={"received_qty": 20}, timeout=15)
        assert r2.status_code == 200, r2.text
        assert r2.json().get("idempotent_replay") is True

        # Capture state after replay
        req_after, se_after, receive_count_after, audit_count_after = asyncio.run(_capture_state())

        # Ledger: exactly one receive entry (no duplicate)
        entries = _get_ledger(admin_token, item_id, dept_id)
        receive_entries = [e for e in entries if e["entry_type"] == "receive"]
        assert len(receive_entries) == 1, f"expected 1 receive entry, got {len(receive_entries)}"

        # received_qty unchanged
        assert req_after.get("received_qty") == req_before.get("received_qty"), \
            f"received_qty changed after replay: {req_before.get('received_qty')} → {req_after.get('received_qty')}"
        # request status unchanged
        assert req_after.get("status") == req_before.get("status"), \
            f"request status changed after replay: {req_before.get('status')} → {req_after.get('status')}"
        # stock balance unchanged
        assert se_after["balance"] == se_before["balance"], \
            f"stock balance changed after replay: {se_before['balance']} → {se_after['balance']}"
        # ledger_version unchanged
        assert se_after["ledger_version"] == se_before["ledger_version"], \
            f"ledger_version changed after replay: {se_before['ledger_version']} → {se_after['ledger_version']}"
        # exactly one receive ledger record
        assert receive_count_after == receive_count_before == 1, \
            f"receive ledger count changed after replay: {receive_count_before} → {receive_count_after}"
        # exactly one receive audit record
        assert audit_count_after == audit_count_before == 1, \
            f"receive audit count changed after replay: {audit_count_before} → {audit_count_after}"


# ---------------------------------------------------------------------------
# Test 5b — receive creates stock entry when none exists
# ---------------------------------------------------------------------------

class TestReceiveMissingStockEntry:
    pytestmark = pytest.mark.integration
    """Correction 2: receive creates stock entry from scratch when none exists."""

    def _setup_request_no_stock(self, token: str) -> dict:
        item = _create_item(token, _unique("T_MISS"))
        dept = _get_er_dept(token)
        # Deliberately do NOT call upsert_stock
        req = _create_request(token, dept["id"], item["id"], qty=15)
        _approve_request(token, req["id"], qty=15)
        _dispatch_request(token, req["id"], qty=15)
        return {"req": req, "item": item, "dept": dept}

    def test_receive_creates_new_stock_entry(self, admin_token):
        ctx = self._setup_request_no_stock(admin_token)
        req_id = ctx["req"]["id"]
        item_id = ctx["item"]["id"]
        dept_id = ctx["dept"]["id"]

        r = requests.post(f"{API}/requests/{req_id}/receive",
                          headers=_h(admin_token), json={"received_qty": 15}, timeout=15)
        assert r.status_code == 200, r.text

        # Stock entry must now exist
        stock_r = requests.get(f"{API}/stock", headers=_h(admin_token),
                               params={"department_id": dept_id, "item_id": item_id}).json()
        entry = next((s for s in stock_r if s["department_id"] == dept_id and s["item_id"] == item_id), None)
        assert entry is not None, "stock entry was not created by receive"
        assert entry["balance"] == 15

        # Exactly one v2 receive ledger record
        entries = _get_ledger(admin_token, item_id, dept_id)
        assert len(entries) == 1, f"expected exactly 1 ledger entry, got {len(entries)}"
        e = entries[0]
        assert e["entry_type"] == "receive"
        assert e["previous_balance"] == 0
        assert e["new_balance"] == 15
        assert e["quantity_change"] == 15
        assert e["sequence_no"] == 1
        assert e["request_id"] == req_id
        assert e["reference_no"] is not None


# ---------------------------------------------------------------------------
# Test 6 — excel_import.commit creates a ledger entry
# ---------------------------------------------------------------------------

class TestExcelImportLedger:
    pytestmark = pytest.mark.integration

    def _make_xlsx(self, internal_code: str, dept_code: str, balance: int) -> bytes:
        """Create a minimal valid .xlsx workbook in memory."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not available in test environment")
        wb = Workbook()
        ws = wb.active
        ws.append([
            "internal_code", "name", "category", "unit",
            "min_level", "critical_threshold", "max_level",
            "department_code", "balance",
        ])
        ws.append([internal_code, f"Import Test {internal_code}", "Other", "PCS",
                   10, 5, 50, dept_code, balance])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_import_creates_ledger_entry(self, admin_token):
        dept = _get_er_dept(admin_token)
        code = _unique("IMP")
        xlsx = self._make_xlsx(code, "ER", 30)

        r = requests.post(
            f"{API}/items/import/commit",
            headers=_h(admin_token),
            data=xlsx,
            timeout=30,
        )
        assert r.status_code == 200, f"import failed: {r.status_code} {r.text}"
        result = r.json()
        assert result["stock_entries_touched"] >= 1, "no stock entries written"

        # Retrieve the item and check ledger
        items = requests.get(f"{API}/items", headers=_h(admin_token)).json()
        item = next((i for i in items if i["internal_code"] == code), None)
        assert item is not None, f"item {code} not found after import"

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        assert len(entries) >= 1, "no ledger entry created by excel import"

        entry = entries[0]
        assert entry["schema_version"] == 2
        assert entry["source"] == "excel_import"
        assert entry["entry_type"] == "physical_count"
        assert entry["new_balance"] == 30
        assert entry["new_balance"] == entry["previous_balance"] + entry["quantity_change"]

    def test_import_same_file_idempotent(self, admin_token):
        dept = _get_er_dept(admin_token)
        code = _unique("IMP2")
        xlsx = self._make_xlsx(code, "ER", 40)

        r1 = requests.post(f"{API}/items/import/commit", headers=_h(admin_token),
                           data=xlsx, timeout=30)
        assert r1.status_code == 200, f"first import failed: {r1.text}"

        r2 = requests.post(f"{API}/items/import/commit", headers=_h(admin_token),
                           data=xlsx, timeout=30)
        assert r2.status_code == 200, f"second import failed: {r2.text}"

        items = requests.get(f"{API}/items", headers=_h(admin_token)).json()
        item = next((i for i in items if i["internal_code"] == code), None)
        assert item is not None

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        import_entries = [e for e in entries if e["source"] == "excel_import"]
        assert len(import_entries) == 1, (
            f"expected exactly 1 import ledger entry after two identical imports, got {len(import_entries)}"
        )

    def test_import_creates_audit_log(self, admin_token):
        """Excel import must write an audit_logs entry with action=excel_stock_import."""
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        dept = _get_er_dept(admin_token)
        code = _unique("IMP3")
        xlsx = self._make_xlsx(code, "ER", 25)

        r = requests.post(
            f"{API}/items/import/commit",
            headers=_h(admin_token),
            data=xlsx,
            timeout=30,
        )
        assert r.status_code == 200, f"import failed: {r.status_code} {r.text}"

        items = requests.get(f"{API}/items", headers=_h(admin_token)).json()
        item = next((i for i in items if i["internal_code"] == code), None)
        assert item is not None, f"item {code} not found after import"

        async def _check_audit():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            _db = c["medstock_test"]
            try:
                log = await _db.audit_logs.find_one(
                    {"action": "excel_stock_import", "entity_id": item["id"]}, {"_id": 0}
                )
                return log
            finally:
                await c.close()
        log = asyncio.run(_check_audit())
        assert log is not None, "no audit_logs entry with action=excel_stock_import found"
        assert log.get("new_value", {}).get("department_id") == dept["id"], \
            f"audit new_value.department_id mismatch: {log.get('new_value')}"
        assert log.get("new_value", {}).get("balance") == 25, \
            f"audit new_value.balance mismatch: {log.get('new_value')}"
        assert log.get("new_value", {}).get("idempotency_key") is not None, \
            f"audit new_value.idempotency_key must be present: {log.get('new_value')}"


# ---------------------------------------------------------------------------
# Test 7 — seed data writes opening_balance ledger entries
# ---------------------------------------------------------------------------

class TestSeedLedger:
    pytestmark = pytest.mark.integration

    def test_seed_items_have_ledger_entries(self, admin_token):
        depts = requests.get(f"{API}/departments", headers=_h(admin_token)).json()
        er = next(d for d in depts if d["code"] == "ER")

        items = requests.get(f"{API}/items", headers=_h(admin_token)).json()
        ett = next((i for i in items if i["internal_code"] == "ETT-CUFF-2"), None)
        if ett is None:
            pytest.skip("Seed data not present")

        entries = _get_ledger(admin_token, ett["id"], er["id"])
        assert entries, "no ledger entries for seeded item"
        opening = [e for e in entries if e["entry_type"] == "opening_balance"]
        assert len(opening) >= 1
        assert opening[0]["actor_type"] == "system"
        assert opening[0]["source"] == "seed"

    def test_seed_idempotency_via_direct_db(self, admin_token, monkeypatch):
        """Run seed() a second time; assert no duplicate stock entries, ledger records, or alerts."""
        import asyncio
        import sys, os as _os
        sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))
        monkeypatch.setenv("ADMIN_EMAIL", _ADMIN_EMAIL)
        monkeypatch.setenv("ADMIN_PASSWORD", _ADMIN_PW)

        async def _run():
            from pymongo import AsyncMongoClient
            from seed import seed as _seed
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0")
            try:
                _db = c["medstock_test"]

                # First seed (ensures data exists)
                await _seed(_db, client=c)

                depts = await _db.departments.find({"code": "ER"}, {"_id": 0}).to_list(1)
                items_list = await _db.items.find({"internal_code": "ETT-CUFF-2"}, {"_id": 0}).to_list(1)
                if not depts or not items_list:
                    return None
                dept_id = depts[0]["id"]
                item_id = items_list[0]["id"]

                # Capture before-snapshot
                before_stock = await _db.stock_entries.find_one(
                    {"department_id": dept_id, "item_id": item_id}, {"_id": 0}
                )
                before_ledger = await _db.stock_transactions.find(
                    {"department_id": dept_id, "item_id": item_id, "schema_version": 2}, {"_id": 0}
                ).to_list(10)
                before_alert_count = await _db.alerts.count_documents(
                    {"department_id": dept_id, "item_id": item_id}
                )

                # Second seed — must be fully idempotent
                await _seed(_db, client=c)

                after_stock = await _db.stock_entries.find_one(
                    {"department_id": dept_id, "item_id": item_id}, {"_id": 0}
                )
                after_ledger = await _db.stock_transactions.find(
                    {"department_id": dept_id, "item_id": item_id, "schema_version": 2}, {"_id": 0}
                ).to_list(10)
                after_alert_count = await _db.alerts.count_documents(
                    {"department_id": dept_id, "item_id": item_id}
                )

                return (before_stock, before_ledger, before_alert_count,
                        after_stock, after_ledger, after_alert_count)
            finally:
                await c.close()
        result = asyncio.run(_run())
        assert result is not None, (
            "seed() failed to create required data: ETT-CUFF-2 item and/or ER department missing "
            "after executing seed() — seed must be idempotent and must always produce required records"
        )

        before_stock, before_ledger, before_alert_count, after_stock, after_ledger, after_alert_count = result

        # Stock entry unchanged
        assert before_stock is not None, "stock entry must exist after first seed"
        assert after_stock is not None, "stock entry must still exist after second seed"
        assert after_stock["balance"] == before_stock["balance"], \
            f"balance changed after second seed: {before_stock['balance']} → {after_stock['balance']}"
        assert after_stock["ledger_version"] == before_stock["ledger_version"], \
            f"ledger_version changed after second seed: {before_stock['ledger_version']} → {after_stock['ledger_version']}"

        # Ledger count unchanged
        assert len(after_ledger) == len(before_ledger), \
            f"ledger count changed after second seed: {len(before_ledger)} → {len(after_ledger)}"

        # Ledger IDs unchanged
        before_ids = {e["id"] for e in before_ledger}
        after_ids = {e["id"] for e in after_ledger}
        assert after_ids == before_ids, f"ledger record IDs changed after second seed"

        # Alert count unchanged
        assert after_alert_count == before_alert_count, \
            f"alert count changed after second seed: {before_alert_count} → {after_alert_count}"

        # Sanity: exactly one ledger record with sequence_no=1
        assert len(after_ledger) == 1, \
            f"expected exactly 1 ledger record for seed item, got {len(after_ledger)}"
        assert after_ledger[0]["sequence_no"] == 1
        assert after_ledger[0]["entry_type"] == "opening_balance"


# ---------------------------------------------------------------------------
# Test 8 — ledger entry fields: schema_version, entry_type, balance invariant
# ---------------------------------------------------------------------------

class TestLedgerEntryFields:
    pytestmark = pytest.mark.integration

    def test_schema_version_is_2(self, admin_token):
        item = _create_item(admin_token, _unique("T8V"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=55)
        entries = _get_ledger(admin_token, item["id"], dept["id"])
        assert entries, "no ledger entries"
        assert all(e["schema_version"] == 2 for e in entries)

    def test_required_fields_present(self, admin_token):
        item = _create_item(admin_token, _unique("T8F"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=33)
        entries = _get_ledger(admin_token, item["id"], dept["id"])
        assert entries
        required = {
            "id", "schema_version", "department_id", "item_id", "entry_type",
            "sequence_no", "previous_balance", "quantity_change", "delta",
            "new_balance", "user_id", "user_name", "actor_type",
            "source", "idempotency_key", "created_at",
        }
        for e in entries:
            missing = required - set(e.keys())
            assert not missing, f"ledger entry missing fields: {missing}"


# ---------------------------------------------------------------------------
# Test 9 — compound unique index: (department_id, item_id, sequence_no) is unique
# ---------------------------------------------------------------------------

class TestLedgerIndexes:
    pytestmark = pytest.mark.integration

    def test_sequence_numbers_are_unique_per_pair(self, admin_token):
        """Verify no duplicate (dept, item, seq_no) tuples across multiple writes."""
        ctx = _create_and_seed(admin_token, balance=100)
        item_id = ctx["item"]["id"]
        dept_id = ctx["dept"]["id"]

        for bal in [90, 80, 70]:
            r = _upsert_stock(admin_token, item_id, dept_id, balance=bal)
            assert r.status_code == 200, r.text

        entries = _get_ledger(admin_token, item_id, dept_id)
        seq_nos = [e["sequence_no"] for e in entries]
        assert len(seq_nos) == len(set(seq_nos)), f"duplicate sequence_no found: {seq_nos}"

    def test_idempotency_key_unique(self, admin_token):
        """Same Idempotency-Key on two different writes is replayed, so only one ledger entry."""
        item = _create_item(admin_token, _unique("T9K"))
        dept = _get_er_dept(admin_token)
        idem_key = _unique("IK9")

        r1 = _upsert_stock(admin_token, item["id"], dept["id"], balance=50, idem_key=idem_key)
        assert r1.status_code == 200, r1.text

        r2 = _upsert_stock(admin_token, item["id"], dept["id"], balance=50, idem_key=idem_key)
        assert r2.status_code == 200, r2.text
        assert r2.json().get("idempotent_replay") is True

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        idem_entries = [e for e in entries if e.get("idempotency_key") == idem_key]
        assert len(idem_entries) == 1, (
            f"expected 1 ledger entry with key {idem_key!r}, got {len(idem_entries)}"
        )


# ---------------------------------------------------------------------------
# Test 10 — immutability: ledger entries must not be modifiable via the API
# ---------------------------------------------------------------------------

class TestLedgerImmutability:
    pytestmark = pytest.mark.integration

    def test_no_update_endpoint(self, admin_token):
        """The API must not expose a PATCH/PUT endpoint for stock transaction entries."""
        ctx = _create_and_seed(admin_token, balance=50)
        entries = _get_ledger(admin_token, ctx["item"]["id"], ctx["dept"]["id"])
        assert entries, "no entries to test immutability on"
        entry_id = entries[0]["id"]

        r = requests.patch(
            f"{API}/stock/transactions/{entry_id}",
            headers=_h(admin_token),
            json={"new_balance": 9999},
            timeout=10,
        )
        assert r.status_code in (404, 405, 422), (
            f"ledger PATCH should not be allowed, got {r.status_code}"
        )

        r = requests.put(
            f"{API}/stock/transactions/{entry_id}",
            headers=_h(admin_token),
            json={"new_balance": 9999},
            timeout=10,
        )
        assert r.status_code in (404, 405, 422), (
            f"ledger PUT should not be allowed, got {r.status_code}"
        )

    def test_no_delete_endpoint(self, admin_token):
        ctx = _create_and_seed(admin_token, balance=50)
        entries = _get_ledger(admin_token, ctx["item"]["id"], ctx["dept"]["id"])
        assert entries
        entry_id = entries[0]["id"]

        r = requests.delete(
            f"{API}/stock/transactions/{entry_id}",
            headers=_h(admin_token),
            timeout=10,
        )
        assert r.status_code in (404, 405, 422), (
            f"ledger DELETE should not be allowed, got {r.status_code}"
        )


# ---------------------------------------------------------------------------
# Test 11 — stock issue ledger entries
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestStockIssueLedger:
    def _issue(self, token, dept_id, item_id, qty):
        return requests.post(f"{API}/stock/issue", headers=_h(token), json={
            "department_id": dept_id,
            "item_id": item_id,
            "quantity": qty,
            "notes": "ledger issue test",
        }, timeout=15)

    def test_issue_creates_one_immutable_record(self, admin_token):
        item = _create_item(admin_token, _unique("TIS1"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], 100)

        r = self._issue(admin_token, dept["id"], item["id"], 10)
        assert r.status_code == 200, r.text

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        issue_entries = [e for e in entries if e["entry_type"] == "issue"]
        assert len(issue_entries) == 1, f"expected 1 issue record, got {len(issue_entries)}"

        e = issue_entries[0]
        # Canonical fields
        required = {"id", "schema_version", "department_id", "item_id", "entry_type", "sequence_no",
                    "previous_balance", "quantity_change", "delta", "new_balance", "status",
                    "user_id", "user_name", "actor_type", "source", "idempotency_key", "created_at", "entry_id"}
        missing = required - set(e.keys())
        assert not missing, f"issue record missing fields: {missing}"
        assert e["quantity_change"] == -10
        assert e["new_balance"] == 90
        assert e["source"] == "stock_issue"
        assert e["actor_type"] == "user"

    def test_issue_second_does_not_need_another_baseline(self, admin_token):
        item = _create_item(admin_token, _unique("TIS2"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], 100)

        r1 = self._issue(admin_token, dept["id"], item["id"], 10)
        assert r1.status_code == 200, r1.text
        r2 = self._issue(admin_token, dept["id"], item["id"], 10)
        assert r2.status_code == 200, r2.text

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        issue_entries = [e for e in entries if e["entry_type"] == "issue"]
        assert len(issue_entries) == 2
        seq_nos = sorted(e["sequence_no"] for e in entries)
        assert seq_nos == list(range(1, len(seq_nos) + 1)), f"sequence gap: {seq_nos}"

    def test_issue_rollback_leaves_no_ledger_entry(self, admin_token):
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        item = _create_item(admin_token, _unique("TIS3"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], 100)

        entries_before = _get_ledger(admin_token, item["id"], dept["id"])
        count_before = len(entries_before)

        async def _snapshot():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                return se["ledger_version"] if se else None
            finally:
                await c.close()
        lv_before = asyncio.run(_snapshot())

        idem_key = _unique("TIS3K")
        headers = _h(admin_token, {"X-Test-Txn-Fail-After": "ledger_insert"})
        r = requests.post(f"{API}/stock/issue", headers=headers, json={
            "department_id": dept["id"], "item_id": item["id"],
            "quantity": 5, "notes": "rollback test",
            "idempotency_key": idem_key,
        }, timeout=15)
        assert r.status_code == 503, f"expected 503, got {r.status_code}"

        entries_after = _get_ledger(admin_token, item["id"], dept["id"])
        assert len(entries_after) == count_before, \
            f"rollback left {len(entries_after) - count_before} extra ledger entries"

        # No issue record
        issue_entries = [e for e in entries_after if e["entry_type"] == "issue"]
        assert issue_entries == [], f"issue ledger record must not exist after rollback: {issue_entries}"

        # Balance and ledger_version unchanged
        stock_r = requests.get(f"{API}/stock", headers=_h(admin_token),
                               params={"department_id": dept["id"], "item_id": item["id"]}).json()
        matching = [s for s in stock_r if s["item_id"] == item["id"] and s["department_id"] == dept["id"]]
        assert matching and matching[0]["balance"] == 100, \
            f"balance must still be 100 after rollback, got {matching}"

        async def _check_after():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                alert = await db.alerts.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                audit = await db.audit_logs.find_one(
                    {"action": {"$in": ["stock_issue", "stock_issue_override"]},
                     "new_value.idempotency_key": idem_key}, {"_id": 0}
                )
                return se, alert, audit
            finally:
                await c.close()
        se, alert, audit = asyncio.run(_check_after())
        assert se["ledger_version"] == lv_before, \
            f"ledger_version must be unchanged: before={lv_before}, after={se['ledger_version']}"
        # No alert created by a rolled-back issue (balance was 100 = available before)
        assert alert is None, f"alert must not exist after rollback: {alert}"
        assert audit is None, f"stock_issue audit must not exist after rollback: {audit}"


# ---------------------------------------------------------------------------
# Test 12 — concurrency
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestConcurrency:
    def test_concurrent_upserts_no_duplicate_sequence(self, admin_token):
        """Two concurrent upserts on the same item — only valid sequences committed."""
        import concurrent.futures
        item = _create_item(admin_token, _unique("TCONC1"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], 100)

        def _do_upsert(balance):
            return _upsert_stock(admin_token, item["id"], dept["id"], balance)

        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as pool:
            futs = [pool.submit(_do_upsert, b) for b in [80, 60]]
            results = [f.result() for f in concurrent.futures.as_completed(futs)]

        # At least one must succeed; none should return 500
        statuses = {r.status_code for r in results}
        assert statuses <= {200, 409, 503}, f"unexpected status codes: {statuses}"

        entries = _get_ledger(admin_token, item["id"], dept["id"])
        seq_nos = [e["sequence_no"] for e in entries]
        assert len(seq_nos) == len(set(seq_nos)), f"duplicate sequence_nos: {seq_nos}"

        # Verify balance invariant for all entries
        for e in entries:
            assert e["previous_balance"] + e["quantity_change"] == e["new_balance"], f"balance invariant violated: {e}"
        # At least one 200
        assert any(r.status_code == 200 for r in results), "at least one upsert must succeed"
        # No unexpected status codes
        assert all(r.status_code in (200, 409, 503) for r in results), f"unexpected codes: {[r.status_code for r in results]}"

    def test_concurrent_receives_on_same_request(self, admin_token):
        """Two concurrent receives of 30 each on a dispatched-40 request.

        Only one of the 30-unit receives can fit (total would be 60 > 40).
        Exactly one must succeed (200) and one must fail (422 or 409/503).
        No duplicate sequence_nos, balance invariant holds.
        """
        import concurrent.futures
        item = _create_item(admin_token, _unique("TCONC2"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], 0)
        req = _create_request(admin_token, dept["id"], item["id"], qty=40)
        _approve_request(admin_token, req["id"], qty=40)
        _dispatch_request(admin_token, req["id"], qty=40)

        idem_keys = [_unique("CRK"), _unique("CRK")]

        def _do_receive(idem_key):
            return requests.post(
                f"{API}/requests/{req['id']}/receive",
                headers=_h(admin_token, {"Idempotency-Key": idem_key}),
                json={"received_qty": 30},
                timeout=15,
            )

        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as pool:
            futs = [pool.submit(_do_receive, k) for k in idem_keys]
            results = [f.result() for f in concurrent.futures.as_completed(futs)]

        statuses = [r.status_code for r in results]
        # Exactly one must succeed; the other must be rejected (over-receive not allowed)
        assert statuses.count(200) == 1, f"expected exactly 1 success, got statuses: {statuses}"
        assert all(s in (200, 409, 422, 503) for s in statuses), f"unexpected statuses: {statuses}"

        # Full DB verification
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        async def _check_full():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                txns = await db.stock_transactions.find(
                    {"department_id": dept["id"], "item_id": item["id"], "schema_version": 2}, {"_id": 0}
                ).sort("sequence_no", 1).to_list(100)
                req_doc = await db.stock_requests.find_one({"id": req["id"]}, {"_id": 0})
                return se, txns, req_doc
            finally:
                await c.close()
        se, txns, req_doc = asyncio.run(_check_full())

        # Exactly one receive record (plus the opening_balance from upsert_stock)
        receive_txns = [t for t in txns if t["entry_type"] == "receive"]
        assert len(receive_txns) == 1, f"expected exactly 1 receive ledger record, got {len(receive_txns)}"

        # No duplicate sequence numbers
        seq_nos = [t["sequence_no"] for t in txns]
        assert len(seq_nos) == len(set(seq_nos)), f"duplicate seq_nos: {seq_nos}"

        # Sequence is continuous
        assert seq_nos == list(range(1, len(seq_nos) + 1)), f"sequence gap: {seq_nos}"

        # Per-record balance arithmetic
        for t in txns:
            assert t["previous_balance"] + t["quantity_change"] == t["new_balance"], \
                f"balance invariant violated in record: {t}"

        # Chained previous_balance: each record's previous_balance == prior record's new_balance
        for i in range(1, len(txns)):
            assert txns[i]["previous_balance"] == txns[i - 1]["new_balance"], \
                f"chain broken at seq {txns[i]['sequence_no']}: prev={txns[i]['previous_balance']} != prior new={txns[i-1]['new_balance']}"

        # stock_entry final balance == last ledger new_balance
        assert se is not None, "stock entry must exist after successful receive"
        assert se["balance"] == txns[-1]["new_balance"], \
            f"stock balance {se['balance']} != last ledger new_balance {txns[-1]['new_balance']}"

        # ledger_version == max(sequence_no)
        max_seq = max(t["sequence_no"] for t in txns)
        assert se["ledger_version"] == max_seq, \
            f"ledger_version={se['ledger_version']} != max(sequence_no)={max_seq}"

        # Stock balance == 30 (exactly one 30-unit receive committed)
        assert se["balance"] == 30, f"expected final stock balance 30, got {se['balance']}"

        # Request received_qty == 30, received_qty <= dispatched_qty
        assert req_doc is not None, "request document not found"
        assert req_doc.get("received_qty") == 30, \
            f"expected request received_qty=30, got {req_doc.get('received_qty')}"
        assert req_doc.get("received_qty", 0) <= req_doc.get("dispatched_qty", 0), \
            f"received_qty {req_doc.get('received_qty')} exceeds dispatched_qty {req_doc.get('dispatched_qty')}"


# ---------------------------------------------------------------------------
# Test 15 — stock issue legacy cutover
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestStockIssueLegacyCutover:

    def test_legacy_cutover_sequence(self, admin_token):
        """Stock entry with no ledger_version gets baseline seq=1, then issue seq=2.

        Uses a fresh API-created item so no shared seeded data is touched.
        The stock entry is inserted directly into the DB without ledger_version
        to simulate a pre-v2 legacy record. No records in stock_transactions
        are ever deleted.
        """
        import asyncio
        import sys, os as _os
        sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        # 1. Create a fresh unique item via the API
        item = _create_item(admin_token, _unique("TLC"))
        dept = _get_er_dept(admin_token)
        item_id = item["id"]
        dept_id = dept["id"]

        # 2. Insert a legacy stock_entry directly (no ledger_version field)
        async def _insert_legacy():
            from pymongo import AsyncMongoClient
            import models
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            _db = c["medstock_test"]
            try:
                entry_id = models._new_id()
                await _db.stock_entries.insert_one({
                    "id": entry_id,
                    "department_id": dept_id,
                    "item_id": item_id,
                    "balance": 50,
                    "status": "available",
                    "last_updated_by": "test",
                    "last_updated_by_name": "Test",
                    "last_updated_at": models._now_iso(),
                    "shortage_start": None,
                    "notes": None,
                    # Intentionally no ledger_version — simulates legacy record
                })
            finally:
                await c.close()
        asyncio.run(_insert_legacy())

        # 3. First issue — should create baseline seq=1 and issue seq=2
        r = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
            "department_id": dept_id,
            "item_id": item_id,
            "quantity": 5,
            "notes": "legacy cutover test",
        }, timeout=15)
        assert r.status_code == 200, r.text

        entries = _get_ledger(admin_token, item_id, dept_id)
        seq_nos = sorted(e["sequence_no"] for e in entries)
        assert 1 in seq_nos, f"baseline seq=1 not found: {seq_nos}"
        assert 2 in seq_nos, f"issue seq=2 not found: {seq_nos}"
        baseline = next(e for e in entries if e["sequence_no"] == 1)
        assert baseline["entry_type"] == "opening_balance"
        issue = next(e for e in entries if e["sequence_no"] == 2)
        assert issue["entry_type"] == "issue"

        # 4. Second issue — no new baseline, seq=3
        r2 = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
            "department_id": dept_id,
            "item_id": item_id,
            "quantity": 5,
            "notes": "legacy cutover test 2",
        }, timeout=15)
        assert r2.status_code == 200, r2.text

        entries2 = _get_ledger(admin_token, item_id, dept_id)
        seq_nos2 = sorted(e["sequence_no"] for e in entries2)
        assert seq_nos2 == [1, 2, 3], f"expected [1,2,3], got {seq_nos2}"
        baselines = [e for e in entries2 if e["entry_type"] == "opening_balance"]
        assert len(baselines) == 1, "should be exactly 1 baseline after 2 issues"


# ---------------------------------------------------------------------------
# Test 16 — idempotency payload conflict returns 409
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestIdempotencyPayloadConflict:

    def test_manual_stock_payload_conflict_returns_409(self, admin_token):
        item = _create_item(admin_token, _unique("TIPC1"))
        dept = _get_er_dept(admin_token)
        idem_key = _unique("TIKPC")

        r1 = _upsert_stock(admin_token, item["id"], dept["id"], balance=50, idem_key=idem_key)
        assert r1.status_code == 200, r1.text

        r2 = _upsert_stock(admin_token, item["id"], dept["id"], balance=99, idem_key=idem_key)
        assert r2.status_code == 409, f"expected 409 for payload conflict, got {r2.status_code}: {r2.text}"

        # Balance must still be 50 (no second mutation)
        entries = _get_ledger(admin_token, item["id"], dept["id"])
        assert all(e["new_balance"] != 99 for e in entries), "balance was mutated despite 409"

    def test_stock_issue_payload_conflict_returns_409(self, admin_token):
        item = _create_item(admin_token, _unique("TIPC2"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=100)
        idem_key = _unique("TIPC2K")

        r1 = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
            "department_id": dept["id"],
            "item_id": item["id"],
            "quantity": 5,
            "idempotency_key": idem_key,
        }, timeout=15)
        assert r1.status_code == 200, r1.text

        r2 = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
            "department_id": dept["id"],
            "item_id": item["id"],
            "quantity": 10,
            "idempotency_key": idem_key,
        }, timeout=15)
        assert r2.status_code == 409, f"expected 409, got {r2.status_code}: {r2.text}"

    def test_receive_payload_conflict_returns_409(self, admin_token):
        item = _create_item(admin_token, _unique("TIPC3"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=0)
        req = _create_request(admin_token, dept["id"], item["id"], qty=40)
        _approve_request(admin_token, req["id"], qty=40)
        _dispatch_request(admin_token, req["id"], qty=40)
        idem_key = _unique("TIPC3K")

        r1 = requests.post(f"{API}/requests/{req['id']}/receive",
                           headers=_h(admin_token, {"Idempotency-Key": idem_key}),
                           json={"received_qty": 20}, timeout=15)
        assert r1.status_code == 200, r1.text

        r2 = requests.post(f"{API}/requests/{req['id']}/receive",
                           headers=_h(admin_token, {"Idempotency-Key": idem_key}),
                           json={"received_qty": 10}, timeout=15)
        assert r2.status_code == 409, f"expected 409, got {r2.status_code}: {r2.text}"

    def test_stock_issue_reference_no_conflict_returns_409(self, admin_token):
        """Reusing an idempotency key with a different reference_no must return 409."""
        item = _create_item(admin_token, _unique("TIPC4"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=100)
        idem_key = _unique("TIPC4K")

        r1 = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
            "department_id": dept["id"],
            "item_id": item["id"],
            "quantity": 5,
            "idempotency_key": idem_key,
            "reference_no": "REF-ORIGINAL",
        }, timeout=15)
        assert r1.status_code == 200, r1.text

        # Same key, same qty, different reference_no — must conflict
        r2 = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
            "department_id": dept["id"],
            "item_id": item["id"],
            "quantity": 5,
            "idempotency_key": idem_key,
            "reference_no": "REF-DIFFERENT",
        }, timeout=15)
        assert r2.status_code == 409, f"expected 409 for reference_no conflict, got {r2.status_code}: {r2.text}"

    def test_stock_issue_approval_id_conflict_returns_409(self, admin_token):
        """Reusing an idempotency key with a different approval_id must return 409."""
        item = _create_item(admin_token, _unique("TIPC5"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=100)
        idem_key = _unique("TIPC5K")

        r1 = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
            "department_id": dept["id"],
            "item_id": item["id"],
            "quantity": 5,
            "idempotency_key": idem_key,
            "approval_id": "APPR-ORIGINAL",
        }, timeout=15)
        assert r1.status_code == 200, r1.text

        r2 = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
            "department_id": dept["id"],
            "item_id": item["id"],
            "quantity": 5,
            "idempotency_key": idem_key,
            "approval_id": "APPR-DIFFERENT",
        }, timeout=15)
        assert r2.status_code == 409, f"expected 409 for approval_id conflict, got {r2.status_code}: {r2.text}"


# ---------------------------------------------------------------------------
# Test 19 — ledger_version invariant
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestVersionInvariant:

    async def _get_lv_and_max_seq(self, dept_id, item_id):
        from pymongo import AsyncMongoClient
        c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
        db = c["medstock_test"]
        try:
            stock = await db.stock_entries.find_one(
                {"department_id": dept_id, "item_id": item_id}, {"_id": 0}
            )
            entries = await db.stock_transactions.find(
                {"department_id": dept_id, "item_id": item_id, "schema_version": 2}, {"_id": 0}
            ).to_list(100)
            lv = stock["ledger_version"] if stock else None
            max_seq = max(e["sequence_no"] for e in entries) if entries else 0
            return lv, max_seq
        finally:
            await c.close()
    def _check(self, dept_id, item_id):
        import asyncio
        return asyncio.run(self._get_lv_and_max_seq(dept_id, item_id))

    def test_manual_stock_path(self, admin_token):
        """ledger_version == max(sequence_no) after two upsert_stock writes."""
        import sys, os as _os
        sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))
        item = _create_item(admin_token, _unique("TVI1"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=100)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=80)
        lv, max_seq = self._check(dept["id"], item["id"])
        assert lv == max_seq, f"manual path: ledger_version={lv} != max(sequence_no)={max_seq}"

    def test_issue_path(self, admin_token):
        """ledger_version == max(sequence_no) after upsert + issue."""
        import sys, os as _os
        sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))
        item = _create_item(admin_token, _unique("TVI2"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=100)
        r = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
            "department_id": dept["id"], "item_id": item["id"],
            "quantity": 10, "notes": "version invariant test",
        }, timeout=15)
        assert r.status_code == 200, r.text
        lv, max_seq = self._check(dept["id"], item["id"])
        assert lv == max_seq, f"issue path: ledger_version={lv} != max(sequence_no)={max_seq}"

    def test_receive_path(self, admin_token):
        """ledger_version == max(sequence_no) after receive_request."""
        import sys, os as _os
        sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))
        item = _create_item(admin_token, _unique("TVI3"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=0)
        req = _create_request(admin_token, dept["id"], item["id"], qty=20)
        _approve_request(admin_token, req["id"], qty=20)
        _dispatch_request(admin_token, req["id"], qty=20)
        r = requests.post(f"{API}/requests/{req['id']}/receive",
                          headers=_h(admin_token), json={"received_qty": 20}, timeout=15)
        assert r.status_code == 200, r.text
        lv, max_seq = self._check(dept["id"], item["id"])
        assert lv == max_seq, f"receive path: ledger_version={lv} != max(sequence_no)={max_seq}"

    def test_excel_path(self, admin_token):
        """ledger_version == max(sequence_no) after excel_import.commit."""
        import sys, os as _os, io
        sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        code = _unique("TVI4")
        wb = Workbook()
        ws = wb.active
        ws.append(["internal_code", "name", "category", "unit",
                   "min_level", "critical_threshold", "max_level", "department_code", "balance"])
        ws.append([code, f"VI Test {code}", "Other", "PCS", 5, 2, 50, "ER", 15])
        buf = io.BytesIO(); wb.save(buf); xlsx = buf.getvalue()

        r = requests.post(f"{API}/items/import/commit", headers=_h(admin_token), data=xlsx, timeout=30)
        assert r.status_code == 200, r.text

        dept = _get_er_dept(admin_token)
        items = requests.get(f"{API}/items", headers=_h(admin_token)).json()
        item = next((i for i in items if i["internal_code"] == code), None)
        assert item is not None
        lv, max_seq = self._check(dept["id"], item["id"])
        assert lv == max_seq, f"excel path: ledger_version={lv} != max(sequence_no)={max_seq}"

    def test_seed_path(self, admin_token, monkeypatch):
        """ledger_version == max(sequence_no) for a seeded item."""
        import asyncio, sys, os as _os
        sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))
        monkeypatch.setenv("ADMIN_EMAIL", _ADMIN_EMAIL)
        monkeypatch.setenv("ADMIN_PASSWORD", _ADMIN_PW)

        async def _ensure_seed():
            from pymongo import AsyncMongoClient
            from seed import seed as _seed
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            try:
                db = c["medstock_test"]
                await _seed(db, client=c)
                depts = await db.departments.find({"code": "ER"}, {"_id": 0}).to_list(1)
                items_list = await db.items.find({"internal_code": "ETT-CUFF-2"}, {"_id": 0}).to_list(1)
                return (depts[0] if depts else None), (items_list[0] if items_list else None)
            finally:
                await c.close()
        er_doc, ett_doc = asyncio.run(_ensure_seed())
        assert er_doc is not None, "ER department not found after seed()"
        assert ett_doc is not None, "ETT-CUFF-2 item not found after seed()"

        lv, max_seq = self._check(er_doc["id"], ett_doc["id"])
        assert lv is not None, "ledger_version must exist for seeded item"
        assert lv == max_seq, f"seed path: ledger_version={lv} != max(sequence_no)={max_seq}"


# ---------------------------------------------------------------------------
# Test 20 — receive_request quantity validation
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestReceiveQuantityValidation:

    def _setup(self, token, qty=40):
        """Create item + request WITHOUT calling _upsert_stock.

        The absence of a pre-existing stock entry means ledger starts empty.
        Any rejected receive must leave ledger, stock, and request state unchanged.
        """
        item = _create_item(token, _unique("TRQ"))
        dept = _get_er_dept(token)
        # Intentionally no _upsert_stock — no opening_balance ledger record
        req = _create_request(token, dept["id"], item["id"], qty=qty)
        _approve_request(token, req["id"], qty=qty)
        _dispatch_request(token, req["id"], qty=qty)
        return item, dept, req

    @staticmethod
    def _capture_receive_audit_count(req_id: str) -> int:
        """Return current count of receive_request audit records for req_id (entity_id field)."""
        import asyncio
        async def _count():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                return await db.audit_logs.count_documents(
                    {"action": "receive_request", "entity_id": req_id}
                )
            finally:
                await c.close()
        return asyncio.run(_count())

    def _assert_no_mutation(self, token, item, dept, req_id, *, before_audit_count: int):
        """Assert ledger, stock, request, and audit state are all unchanged after a rejected receive.

        before_audit_count must be captured before the rejected request so the comparison
        is before/after rather than an unchecked assertion of zero.
        """
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        ledger_entries = _get_ledger(token, item["id"], dept["id"])
        assert ledger_entries == [], \
            f"no ledger entry should exist, got {len(ledger_entries)}"

        stock_r = requests.get(f"{API}/stock", headers=_h(token),
                               params={"department_id": dept["id"], "item_id": item["id"]}).json()
        matching = [s for s in stock_r if s["item_id"] == item["id"] and s["department_id"] == dept["id"]]
        assert not matching, f"stock entry must not exist after rejected receive: {matching}"

        req_state = _get_request(token, req_id)
        assert req_state.get("received_qty", 0) == 0, \
            f"request received_qty must still be 0, got {req_state.get('received_qty')}"
        assert req_state["status"] == "dispatched", \
            f"request status must still be 'dispatched', got {req_state['status']}"

        audit_count_after = self._capture_receive_audit_count(req_id)
        assert audit_count_after == before_audit_count, (
            f"receive_request audit count changed after rejected receive: "
            f"{before_audit_count} → {audit_count_after}"
        )

    def test_zero_qty_rejected(self, admin_token):
        item, dept, req = self._setup(admin_token)
        before_audit_count = self._capture_receive_audit_count(req["id"])
        r = requests.post(f"{API}/requests/{req['id']}/receive",
                          headers=_h(admin_token), json={"received_qty": 0}, timeout=15)
        assert r.status_code == 422, f"expected 422 for zero qty, got {r.status_code}: {r.text}"
        self._assert_no_mutation(admin_token, item, dept, req["id"],
                                 before_audit_count=before_audit_count)

    def test_negative_qty_rejected(self, admin_token):
        item, dept, req = self._setup(admin_token)
        before_audit_count = self._capture_receive_audit_count(req["id"])
        r = requests.post(f"{API}/requests/{req['id']}/receive",
                          headers=_h(admin_token), json={"received_qty": -5}, timeout=15)
        assert r.status_code == 422, f"expected 422 for negative qty, got {r.status_code}: {r.text}"
        self._assert_no_mutation(admin_token, item, dept, req["id"],
                                 before_audit_count=before_audit_count)

    def test_excess_qty_rejected(self, admin_token):
        """Receiving more than dispatched remaining must return 422."""
        item, dept, req = self._setup(admin_token, qty=10)
        before_audit_count = self._capture_receive_audit_count(req["id"])
        r = requests.post(f"{API}/requests/{req['id']}/receive",
                          headers=_h(admin_token), json={"received_qty": 99}, timeout=15)
        assert r.status_code == 422, f"expected 422 for excess qty, got {r.status_code}: {r.text}"
        self._assert_no_mutation(admin_token, item, dept, req["id"],
                                 before_audit_count=before_audit_count)


# ---------------------------------------------------------------------------
# Test 21 — excel import rollback
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestExcelRollback:

    def _make_xlsx_update(self, internal_code: str, dept_code: str, balance: int) -> bytes:
        """Build xlsx referencing an existing internal_code so the plan puts it in to_update."""
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not available in test environment")
        import io
        wb = Workbook()
        ws = wb.active
        ws.append(["internal_code", "name", "category", "unit",
                   "min_level", "critical_threshold", "max_level", "department_code", "balance"])
        ws.append([internal_code, f"Rollback Test {internal_code}", "Other", "PCS",
                   5, 2, 50, dept_code, balance])
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def test_rollback_at_excel_stock_update_leaves_no_entry(self, admin_token):
        """X-Test-Fail-After: excel_stock_update must roll back completely — no residue in any collection."""
        import asyncio, hashlib, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        dept = _get_er_dept(admin_token)
        # Pre-create item via API so item_id is known unconditionally
        item = _create_item(admin_token, _unique("XLRB"))
        item_id = item["id"]

        # Capture full item document before import (all Excel-modifiable fields)
        async def _capture_item():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                return await db.items.find_one({"id": item_id}, {"_id": 0})
            finally:
                await c.close()
        item_doc_before = asyncio.run(_capture_item())
        assert item_doc_before is not None, "item must exist before import"

        xlsx = self._make_xlsx_update(item["internal_code"], "ER", 20)
        expected_sha = hashlib.sha256(xlsx).hexdigest()

        r = requests.post(
            f"{API}/items/import/commit",
            headers=_h(admin_token, {"X-Test-Fail-After": "excel_stock_update"}),
            data=xlsx,
            timeout=30,
        )
        assert r.status_code == 503, \
            f"expected 503 for injected rollback, got {r.status_code}: {r.text}"

        async def _check_all():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                item_doc_after = await db.items.find_one({"id": item_id}, {"_id": 0})
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item_id}, {"_id": 0}
                )
                physical_count = await db.stock_transactions.find_one(
                    {"department_id": dept["id"], "item_id": item_id,
                     "schema_version": 2, "entry_type": "physical_count"}, {"_id": 0}
                )
                baseline = await db.stock_transactions.find_one(
                    {"department_id": dept["id"], "item_id": item_id,
                     "schema_version": 2, "entry_type": "opening_balance"}, {"_id": 0}
                )
                row_audit = await db.audit_logs.find_one(
                    {"action": "excel_stock_import", "entity_id": item_id}, {"_id": 0}
                )
                summary_audit = await db.audit_logs.find_one(
                    {"action": "import_excel", "new_value.workbook_sha256": expected_sha}, {"_id": 0}
                )
                alert = await db.alerts.find_one(
                    {"department_id": dept["id"], "item_id": item_id}, {"_id": 0}
                )
                return item_doc_after, se, physical_count, baseline, row_audit, summary_audit, alert
            finally:
                await c.close()
        item_doc_after, se, physical_count, baseline, row_audit, summary_audit, alert = asyncio.run(_check_all())

        # Item must still exist and all Excel-modifiable fields must be unchanged
        assert item_doc_after is not None, "item must still exist after rollback"
        _excel_fields = ["barcode", "name_en", "name_ar", "category", "unit",
                         "min_level", "critical_threshold", "max_level", "updated_at"]
        for field in _excel_fields:
            assert item_doc_after.get(field) == item_doc_before.get(field), (
                f"item.{field} must not change after rollback: "
                f"{item_doc_before.get(field)!r} → {item_doc_after.get(field)!r}"
            )

        assert se is None, f"stock_entry must not exist after rollback: {se}"
        assert physical_count is None, \
            f"physical_count ledger record must not exist after rollback: {physical_count}"
        assert baseline is None, \
            f"opening_balance baseline must not exist after rollback: {baseline}"
        assert row_audit is None, \
            f"row-level audit must not exist after rollback: {row_audit}"
        assert summary_audit is None, \
            f"import_excel summary audit (sha={expected_sha}) must not exist after rollback: {summary_audit}"
        assert alert is None, f"alert must not exist after rollback: {alert}"


# ---------------------------------------------------------------------------
# Test 22 — Excel item-identity conflict returns 409
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestExcelItemIdentityConflict:

    def test_conflicting_item_identity_returns_409(self, admin_token):
        """Idempotency key reused for a different item_id must return 409 (ExcelWriteConflict)."""
        import asyncio, hashlib, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))
        try:
            from openpyxl import Workbook
        except ImportError:
            pytest.skip("openpyxl not available")

        dept = _get_er_dept(admin_token)

        # Create item A (will be targeted by the xlsx)
        item_a = _create_item(admin_token, _unique("XICA"))
        # Create item B (will be planted in the conflicting ledger fixture)
        item_b = _create_item(admin_token, _unique("XICB"))

        # Build xlsx targeting item A's internal_code with balance=30
        wb = Workbook()
        ws = wb.active
        ws.append(["internal_code", "name", "category", "unit",
                   "min_level", "critical_threshold", "max_level", "department_code", "balance"])
        ws.append([item_a["internal_code"], f"XIC Test {item_a['internal_code']}", "Other", "PCS",
                   5, 2, 50, "ER", 30])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        # Compute the idempotency key that commit() will use for row 2
        from ledger import workbook_row_idempotency_key
        idem_key = workbook_row_idempotency_key(xlsx_bytes, 2)

        # Plant a conflicting ledger record using item B's id but the same idem_key
        async def _plant_conflict():
            from pymongo import AsyncMongoClient
            from models import _new_id, _now_iso
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                await db.stock_transactions.insert_one({
                    "id": _new_id(),
                    "schema_version": 2,
                    "source": "excel_import",
                    "entry_type": "physical_count",
                    "department_id": dept["id"],
                    "item_id": item_b["id"],   # deliberately item B
                    "new_balance": 30,
                    "previous_balance": 0,
                    "quantity_change": 30,
                    "delta": 30,
                    "sequence_no": 1,
                    "idempotency_key": idem_key,
                    "status": "available",
                    "user_id": "test",
                    "user_name": "Test",
                    "actor_type": "user",
                    "entry_id": _new_id(),
                    "created_at": _now_iso(),
                })
            finally:
                await c.close()
        asyncio.run(_plant_conflict())

        # Capture full item A document BEFORE submitting the workbook
        async def _fetch_item_a():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                return await db.items.find_one({"id": item_a["id"]}, {"_id": 0})
            finally:
                await c.close()
        item_a_before = asyncio.run(_fetch_item_a())
        assert item_a_before is not None, "item A must exist in DB before import"

        # Submit xlsx — must return 409 because prior record has item_b["id"] != item_a["id"]
        r = requests.post(
            f"{API}/items/import/commit",
            headers=_h(admin_token),
            data=xlsx_bytes,
            timeout=30,
        )
        assert r.status_code == 409, \
            f"expected 409 for item-identity conflict, got {r.status_code}: {r.text}"

        # Fetch item A again after 409 and compare all Excel-modifiable fields
        async def _verify_no_mutation():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                item_a_after = await db.items.find_one({"id": item_a["id"]}, {"_id": 0})
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item_a["id"]}, {"_id": 0}
                )
                physical_count = await db.stock_transactions.find_one(
                    {"department_id": dept["id"], "item_id": item_a["id"],
                     "schema_version": 2, "entry_type": "physical_count"}, {"_id": 0}
                )
                row_audit = await db.audit_logs.find_one(
                    {"action": "excel_stock_import", "entity_id": item_a["id"]}, {"_id": 0}
                )
                summary_audit = await db.audit_logs.find_one(
                    {"action": "import_excel",
                     "new_value.workbook_sha256": hashlib.sha256(xlsx_bytes).hexdigest()}, {"_id": 0}
                )
                return item_a_after, se, physical_count, row_audit, summary_audit
            finally:
                await c.close()
        item_a_after, se, physical_count, row_audit, summary_audit = asyncio.run(_verify_no_mutation())

        # Item A document must be completely unchanged
        assert item_a_after is not None, "item A must still exist in DB after 409"
        _compare_fields = [
            "id", "internal_code", "barcode", "name_en", "name_ar",
            "category", "unit", "min_level", "critical_threshold", "max_level", "updated_at",
        ]
        for field in _compare_fields:
            assert item_a_after.get(field) == item_a_before.get(field), (
                f"item A field '{field}' must be unchanged after 409: "
                f"{item_a_before.get(field)!r} → {item_a_after.get(field)!r}"
            )

        assert se is None, f"stock entry for item A must not exist after 409: {se}"
        assert physical_count is None, \
            f"physical_count ledger for item A must not exist after 409: {physical_count}"
        assert row_audit is None, \
            f"row-level audit for item A must not exist after 409: {row_audit}"
        assert summary_audit is None, \
            f"import_excel summary audit must not exist after 409: {summary_audit}"


# ---------------------------------------------------------------------------
# Test 23 — reserved idempotency-key prefix rejection
# ---------------------------------------------------------------------------

@pytest.mark.integration
class TestReservedPrefixRejection:
    """All three reserved prefixes (baseline:, excel:, seed:) must be rejected with HTTP 422
    by upsert_stock, stock_issue, and receive_request. No state mutation may occur."""

    _PREFIXES = ["baseline:", "excel:", "seed:"]

    def test_upsert_stock_rejects_reserved_prefixes(self, admin_token):
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        item = _create_item(admin_token, _unique("TRPU"))
        dept = _get_er_dept(admin_token)

        # Pre-generate all three reserved keys so the audit query can correlate exactly
        reserved_keys = [f"{p}{_unique()}" for p in self._PREFIXES]

        async def _snapshot():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                ledger_count = await db.stock_transactions.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"], "schema_version": 2}
                )
                audit_count = await db.audit_logs.count_documents(
                    {"action": "upsert_stock",
                     "new_value.idempotency_key": {"$in": reserved_keys}}
                )
                alert_count = await db.alerts.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"]}
                )
                return se, ledger_count, audit_count, alert_count
            finally:
                await c.close()
        se_before, ledger_count_before, audit_count_before, alert_count_before = asyncio.run(_snapshot())

        for idem_key in reserved_keys:
            r = _upsert_stock(admin_token, item["id"], dept["id"], balance=42, idem_key=idem_key)
            assert r.status_code == 422, \
                f"upsert_stock with key {idem_key!r} must return 422, got {r.status_code}: {r.text}"

        se_after, ledger_count_after, audit_count_after, alert_count_after = asyncio.run(_snapshot())

        # stock entry: unchanged (absent before → still absent after)
        if se_before is None:
            assert se_after is None, \
                f"stock entry must still be absent after reserved-prefix rejections: {se_after}"
        else:
            assert se_after is not None, "stock entry disappeared after reserved-prefix rejections"
            assert se_after["balance"] == se_before["balance"], \
                f"balance changed: {se_before['balance']} → {se_after['balance']}"
            assert se_after.get("ledger_version") == se_before.get("ledger_version"), \
                f"ledger_version changed: {se_before.get('ledger_version')} → {se_after.get('ledger_version')}"
        assert ledger_count_after == ledger_count_before, \
            f"ledger count changed: {ledger_count_before} → {ledger_count_after}"
        assert audit_count_after == audit_count_before, \
            f"upsert_stock audit count changed: {audit_count_before} → {audit_count_after}"
        assert alert_count_after == alert_count_before, \
            f"alert count changed: {alert_count_before} → {alert_count_after}"

    def test_stock_issue_rejects_reserved_prefixes(self, admin_token):
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        item = _create_item(admin_token, _unique("TRPI"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=100)

        # Pre-generate all three reserved keys so the audit query can correlate exactly
        reserved_keys = [f"{p}{_unique()}" for p in self._PREFIXES]

        async def _snapshot():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                ledger_count = await db.stock_transactions.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"], "schema_version": 2}
                )
                audit_count = await db.audit_logs.count_documents(
                    {"action": {"$in": ["stock_issue", "stock_issue_override"]},
                     "new_value.idempotency_key": {"$in": reserved_keys}}
                )
                alert_count = await db.alerts.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"]}
                )
                return se, ledger_count, audit_count, alert_count
            finally:
                await c.close()
        se_before, ledger_count_before, audit_count_before, alert_count_before = asyncio.run(_snapshot())

        for idem_key in reserved_keys:
            r = requests.post(f"{API}/stock/issue", headers=_h(admin_token), json={
                "department_id": dept["id"],
                "item_id": item["id"],
                "quantity": 5,
                "idempotency_key": idem_key,
            }, timeout=15)
            assert r.status_code == 422, \
                f"stock_issue with key {idem_key!r} must return 422, got {r.status_code}: {r.text}"

        se_after, ledger_count_after, audit_count_after, alert_count_after = asyncio.run(_snapshot())

        assert se_after is not None, "stock entry must still exist after reserved-prefix rejections"
        assert se_after["balance"] == se_before["balance"], \
            f"balance changed: {se_before['balance']} → {se_after['balance']}"
        assert se_after.get("ledger_version") == se_before.get("ledger_version"), \
            f"ledger_version changed: {se_before.get('ledger_version')} → {se_after.get('ledger_version')}"
        assert ledger_count_after == ledger_count_before, \
            f"ledger count changed: {ledger_count_before} → {ledger_count_after}"
        assert audit_count_after == audit_count_before, \
            f"stock_issue audit count changed: {audit_count_before} → {audit_count_after}"
        assert alert_count_after == alert_count_before, \
            f"alert count changed: {alert_count_before} → {alert_count_after}"

    def test_receive_rejects_reserved_prefixes(self, admin_token):
        import asyncio, sys as _sys, os as _os
        _sys.path.insert(0, _os.path.join(_os.path.dirname(__file__), ".."))

        item = _create_item(admin_token, _unique("TRPR"))
        dept = _get_er_dept(admin_token)
        _upsert_stock(admin_token, item["id"], dept["id"], balance=0)
        req = _create_request(admin_token, dept["id"], item["id"], qty=30)
        _approve_request(admin_token, req["id"], qty=30)
        _dispatch_request(admin_token, req["id"], qty=30)

        async def _snapshot():
            from pymongo import AsyncMongoClient
            c = AsyncMongoClient("mongodb://mongo:27017/?replicaSet=rs0", serverSelectionTimeoutMS=5000)
            db = c["medstock_test"]
            try:
                se = await db.stock_entries.find_one(
                    {"department_id": dept["id"], "item_id": item["id"]}, {"_id": 0}
                )
                req_doc = await db.stock_requests.find_one({"id": req["id"]}, {"_id": 0})
                receive_count = await db.stock_transactions.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"],
                     "schema_version": 2, "entry_type": "receive"}
                )
                audit_count = await db.audit_logs.count_documents(
                    {"action": "receive_request", "entity_id": req["id"]}
                )
                alert_count = await db.alerts.count_documents(
                    {"department_id": dept["id"], "item_id": item["id"]}
                )
                return se, req_doc, receive_count, audit_count, alert_count
            finally:
                await c.close()
        se_before, req_before, receive_count_before, audit_count_before, alert_count_before = asyncio.run(_snapshot())

        for prefix in self._PREFIXES:
            idem_key = f"{prefix}{_unique()}"
            r = requests.post(
                f"{API}/requests/{req['id']}/receive",
                headers=_h(admin_token, {"Idempotency-Key": idem_key}),
                json={"received_qty": 10},
                timeout=15,
            )
            assert r.status_code == 422, \
                f"receive with prefix {prefix!r} must return 422, got {r.status_code}: {r.text}"

        se_after, req_after, receive_count_after, audit_count_after, alert_count_after = asyncio.run(_snapshot())

        bal_before = se_before["balance"] if se_before else 0
        bal_after = se_after["balance"] if se_after else 0
        assert bal_after == bal_before, \
            f"stock balance changed: {bal_before} → {bal_after}"
        lv_before = se_before.get("ledger_version", 0) if se_before else 0
        lv_after = se_after.get("ledger_version", 0) if se_after else 0
        assert lv_after == lv_before, \
            f"ledger_version changed: {lv_before} → {lv_after}"
        assert req_after.get("received_qty", 0) == req_before.get("received_qty", 0), \
            f"received_qty changed: {req_before.get('received_qty')} → {req_after.get('received_qty')}"
        assert req_after.get("status") == req_before.get("status"), \
            f"request status changed: {req_before.get('status')} → {req_after.get('status')}"
        assert receive_count_after == receive_count_before, \
            f"receive ledger count changed: {receive_count_before} → {receive_count_after}"
        assert audit_count_after == audit_count_before, \
            f"receive_request audit count changed: {audit_count_before} → {audit_count_after}"
        assert alert_count_after == alert_count_before, \
            f"alert count changed: {alert_count_before} → {alert_count_after}"
