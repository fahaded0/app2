"""Excel (.xlsx) and PDF builders for the 10 formal reports.

Each report follows the same metadata header:
    - Report name
    - Period (optional)
    - Issue date (auto)
    - Extracted by (current user)
    - Record count
    - System notes (optional)
"""
import io
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
)


# ---------- Excel ----------
_HEAD_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
_HEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_META_FONT = Font(name="Calibri", bold=True, size=10)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="0F172A")
_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def build_excel(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    meta: dict,
    sheet_name: str = "Report",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    # --- Title bar ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # --- Metadata block (4 rows: period / issue date / extracted by / record count) ---
    meta_rows = [
        ("Period",        meta.get("period", "—")),
        ("Issue Date",    meta.get("issue_date", datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC"))),
        ("Extracted By",  meta.get("extracted_by", "—")),
        ("Record Count",  str(meta.get("count", len(rows)))),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=k).font = _META_FONT
        ws.cell(row=i, column=2, value=v)
    if meta.get("notes"):
        ws.cell(row=6, column=1, value="System notes").font = _META_FONT
        ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=len(headers))
        ws.cell(row=6, column=2, value=meta["notes"])
        header_start_row = 8
    else:
        header_start_row = 7

    # --- Header row ---
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_start_row, column=c, value=h)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _BORDER
    ws.row_dimensions[header_start_row].height = 22

    # --- Data ---
    for r_idx, row in enumerate(rows, start=header_start_row + 1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"

    # --- Auto column widths ---
    for c in range(1, len(headers) + 1):
        max_len = len(str(headers[c - 1])) + 2
        for r in rows:
            v = r[c - 1] if c - 1 < len(r) else ""
            max_len = max(max_len, min(60, len(str(v)) + 2))
        ws.column_dimensions[get_column_letter(c)].width = max_len

    ws.freeze_panes = ws.cell(row=header_start_row + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------- PDF ----------
def build_pdf(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    meta: dict,
) -> bytes:
    buf = io.BytesIO()
    page_size = landscape(A4) if len(headers) > 5 else A4
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"], alignment=0,
        textColor=colors.HexColor("#0F172A"), fontSize=16, leading=20,
    )
    meta_style = ParagraphStyle(
        "Meta", parent=styles["Normal"], fontSize=9, leading=12,
        textColor=colors.HexColor("#475569"),
    )

    flow: list[Any] = []
    flow.append(Paragraph(title, title_style))
    flow.append(Paragraph("Critical Medical Stock Monitoring System", meta_style))
    flow.append(Spacer(1, 6))

    meta_rows = [
        ["Period:",       meta.get("period", "—"),
         "Issue Date:",   meta.get("issue_date", datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC"))],
        ["Extracted By:", meta.get("extracted_by", "—"),
         "Records:",      str(meta.get("count", len(rows)))],
    ]
    meta_tbl = Table(meta_rows, colWidths=[28 * mm, 70 * mm, 28 * mm, 60 * mm])
    meta_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#0F172A")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow.append(meta_tbl)

    if meta.get("notes"):
        flow.append(Spacer(1, 4))
        flow.append(Paragraph(f"<b>System notes:</b> {meta['notes']}", meta_style))

    flow.append(Spacer(1, 10))

    # --- Data table ---
    table_data = [headers] + [
        ["" if v is None else str(v) for v in row] for row in rows
    ]
    avail_width = page_size[0] - 30 * mm
    if not headers:
        col_widths = []
    else:
        col_widths = [avail_width / len(headers)] * len(headers)
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
        ("ALIGN", (0, 0), (-1, 0), "LEFT"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow.append(tbl)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#94A3B8"))
        canvas.drawString(15 * mm, 8 * mm,
                          f"Critical Medical Stock Monitor · Confidential · {title}")
        canvas.drawRightString(page_size[0] - 15 * mm, 8 * mm,
                               f"Page {doc_.page}")
        canvas.restoreState()

    doc.build(flow, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    return buf.getvalue()
